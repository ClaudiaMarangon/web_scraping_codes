#!/usr/bin/env python2
# -*- coding: utf-8 -*-
"""
Created on Tue Mar  6 11:58:02 2018

@author: claudiamarangon
"""

#!/usr/bin/env python2
# -*- coding: utf-8 -*-
"""
Created on Fri Feb 23 13:58:08 2018

@author: claudiamarangon
"""
import os
import requests
import openpyxl
import smtplib
import time

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
os.environ["PATH"] += "/usr/local/bin"

"""OPEN EXCEL FILE"""
wb = openpyxl.load_workbook('Word_Count.xlsx')
b1 = wb['Burocrazia']
b1['A1']= 'Year'
b1['B1']= 'Month'
b1['C1']= 'word_freq'
b1['D1']= 'word_freq_tit'
b1['E1']= 'word_freq_1_p'
b1['F1']= 'word_freq_tit_1_p'
wb.save('Word_Count.xlsx')
"""SELECT TIME INFO + SEARCH FOR WORD"""
chromedriver = '/usr/local/bin/chromedriver'
browser = webdriver.Chrome(executable_path=chromedriver)

try:
    year = 1876

    row_n = 2
    browser.get('INSERT URL')
        
    wait = WebDriverWait(browser, 60)
    element = wait.until(EC.element_to_be_clickable((By.ID, 'txbSearch')))
        
    search_elem = browser.find_element_by_id('txbSearch')
    search_elem.send_keys('burocrazia')

    while year<=2017:
        bis1 = year%4    
        if year == 1876:
            month = 3
            month_in = '03'
        else:
            month = 1
            month_in = '01'
        while month<= 12:
            
            
            guided_s = browser.find_element_by_xpath("//div[@id='mainsearch']/div/div[1]/table/body/tr/td[2]/div[2]/table/body/tr[3]/td[1]")
            guided_s.click()
                
            text = browser.find_element_by_xpath("//div[@id='mainsearch']/div/div[1]/table/body/tr/td[2]/div[2]/table/body/tr[3]/td[1]/div/div/ul/li[2]")
            text.click()
               
            if month== 4 or month == 6 or month == 9 or month == 11:
                date_end = 30
            elif month == 2 and (bis1==0 and year!=1900):
                date_end = 29
            elif month == 2 and (bis1!=0 or year==1900):
                date_end = 28
            else:
                date_end = 31
        

            beg_date = browser.find_element_by_id('date_start')
            
            n = 1
            while n<=14:
                beg_date.send_keys(Keys.BACK_SPACE)
                n = n+1
        
            if month<10:
                beg_date.send_keys('01/0'+str(month)+'/'+str(year))
            else:
                beg_date.send_keys('01/'+str(month)+'/'+str(year))
        
            end_date = browser.find_element_by_id('date_end')
        
            n = 1
            while n<=14:
                end_date.send_keys(Keys.BACK_SPACE)
                n = n+1

        
            if month<10:
                end_date.send_keys(str(date_end)+'/0'+str(month)+'/'+str(year))
            else:
                end_date.send_keys(str(date_end)+'/'+str(month)+'/'+str(year))
              
            time.sleep(.5)
            search_elem.send_keys(Keys.ENTER)
        
            """GET ELEMENT"""
            wait = WebDriverWait(browser, 60)
            element = wait.until(EC.text_to_be_present_in_element((By.ID, 'risultaticount'), "trov"))
            find_elem = browser.find_element_by_id('risultaticount')
            find_u = find_elem.text
            find_s = find_u.encode("utf-8")
            if find_s == 'Nessun documento trovato':
                count_burocrazia = 0
            else:
                count = ''
                k = 9
                while find_s[k]!=' ':
                    count = count + find_s[k]
                    k = k + 1
                count_burocrazia = int(count)
            
            back = browser.find_element_by_id('home')
            back.click()
            
        
        
            """SEARCH TITLE ELEMENT"""
            
            
            element = wait.until(EC.element_to_be_clickable((By.ID, 'txbSearch')))
                
            guided_s = browser.find_element_by_xpath("//div[@id='mainsearch']/div/div[1]/table/body/tr/td[2]/div[2]/table/body/tr[3]/td[1]")
            guided_s.click()
                
            tit = browser.find_element_by_xpath("//div[@id='mainsearch']/div/div[1]/table/body/tr/td[2]/div[2]/table/body/tr[3]/td[1]/div/div/ul/li[5]")
            tit.click()
                
        
            beg_date = browser.find_element_by_id('date_start')
            n = 1
            while n<=14:
                beg_date.send_keys(Keys.BACK_SPACE)
                n = n+1            
            if month<10:
                beg_date.send_keys(Keys.BACK_SPACE, '01/0'+str(month)+'/'+str(year))
            else:
                beg_date.send_keys(Keys.BACK_SPACE, '01/'+str(month)+'/'+str(year))
        
            end_date = browser.find_element_by_id('date_end')
        
            n = 1
            while n<=14:
                end_date.send_keys(Keys.BACK_SPACE)
                n = n+1
    
        
            if month<10:
                end_date.send_keys(Keys.BACK_SPACE, str(date_end)+'/0'+str(month)+'/'+str(year))
            else:
                end_date.send_keys(Keys.BACK_SPACE, str(date_end)+'/'+str(month)+'/'+str(year))
            
        
            time.sleep(.5)
            search_elem.send_keys(Keys.ENTER)
            
                        
            back = browser.find_element_by_id('home')
            back.click()
            
            wait = WebDriverWait(browser, 60)
            element = wait.until(EC.text_to_be_present_in_element((By.ID, 'risultaticount'), "trov"))
            find_elem = browser.find_element_by_id('risultaticount')
            find_u = find_elem.text
            find_s = find_u.encode("utf-8")
            if find_s == 'Nessun documento trovato':
                count_tit = 0
            else:
                count = ''
                k = 9
                while find_s[k]!=' ':
                    count = count + find_s[k]
                    k = k + 1
                count_tit = int(count)
            
            back = browser.find_element_by_id('home')
            back.click()
    
            """SEARCH FIRST PAGE ELEMENT"""
            
            if count_burocrazia == 0:
                count_burocrazia_1_p = 0
            else:
                element = wait.until(EC.element_to_be_clickable((By.ID, 'txbSearch')))
                
                
                guided_s = browser.find_element_by_xpath("//div[@id='mainsearch']/div/div[1]/table/body/tr/td[2]/div[2]/table/body/tr[3]/td[1]/div")
                guided_s.click()
                
                text = browser.find_element_by_xpath("//div[@id='mainsearch']/div/div[1]/table/body/tr/td[2]/div[2]/table/body/tr[3]/td[1]/div/div/ul/li[2]")
                text.click()
                
                
                page = browser.find_element_by_xpath("//div[@id='mainsearch']/div/div[1]/table/body/tr/td[2]/div[2]/table/body/tr[3]/td[2]/div")
                page.click()
                first_p = browser.find_element_by_xpath("//div[@id='mainsearch']/div/div[1]/table/body/tr/td[2]/div[2]/table/body/tr[3]/td[2]/div/div/ul/li[1]")
                first_p.click()
        
                beg_date = browser.find_element_by_id('date_start')
                
                n = 1
                while n<=14:
                    beg_date.send_keys(Keys.BACK_SPACE)
                    n = n+1
                
                if month<10:
                    beg_date.send_keys(Keys.BACK_SPACE, '01/0'+str(month)+'/'+str(year))
                else:
                    beg_date.send_keys(Keys.BACK_SPACE, '01/'+str(month)+'/'+str(year))
        
                end_date = browser.find_element_by_id('date_end')
        
                n = 1
                while n<=14:
                    end_date.send_keys(Keys.BACK_SPACE)
                    n = n+1
    
        
                if month<10:
                    end_date.send_keys(Keys.BACK_SPACE, str(date_end)+'/0'+str(month)+'/'+str(year))
                else:
                    end_date.send_keys(Keys.BACK_SPACE, str(date_end)+'/'+str(month)+'/'+str(year))
            
        
                time.sleep(.5)
                search_elem.send_keys(Keys.ENTER)
    
            
            
                wait = WebDriverWait(browser, 60)
                element = wait.until(EC.text_to_be_present_in_element((By.ID, 'risultaticount'), "trov"))
                find_elem = browser.find_element_by_id('risultaticount')
                find_u = find_elem.text
                find_s = find_u.encode("utf-8")
                if find_s == 'Nessun documento trovato':
                    count_burocrazia = 0
                else:
                    count = ''
                    k = 9
                    while find_s[k]!=' ':
                        count = count + find_s[k]
                        k = k + 1
                        count_burocrazia = int(count)
            
                back = browser.find_element_by_id('home')
                back.click()
                
        
                time.sleep(.5)                
                page = browser.find_element_by_xpath("//div[@id='mainsearch']/div/div[1]/table/body/tr/td[2]/div[2]/table/body/tr[3]/td[2]/div")
                page.click()
                first_p = browser.find_element_by_xpath("//div[@id='mainsearch']/div/div[1]/table/body/tr/td[2]/div[2]/table/body/tr[3]/td[2]/div/div/ul/li[1]")
                first_p.click()

            """SEARCH FIRST PAGE TITLE ELEMENT"""
            
            if count_tit == 0:
                count_tit_1_p = 0
            else:
                element = wait.until(EC.element_to_be_clickable((By.ID, 'txbSearch')))
                
                
                guided_s = browser.find_element_by_xpath("//div[@id='mainsearch']/div/div[1]/table/body/tr/td[2]/div[2]/table/body/tr[3]/td[1]/div")
                guided_s.click()
                
                tit = browser.find_element_by_xpath("//div[@id='mainsearch']/div/div[1]/table/body/tr/td[2]/div[2]/table/body/tr[3]/td[1]/div/div/ul/li[5]")
                tit.click()
                
                
                page = browser.find_element_by_xpath("//div[@id='mainsearch']/div/div[1]/table/body/tr/td[2]/div[2]/table/body/tr[3]/td[2]/div")
                page.click()
                first_p = browser.find_element_by_xpath("//div[@id='mainsearch']/div/div[1]/table/body/tr/td[2]/div[2]/table/body/tr[3]/td[2]/div/div/ul/li[1]")
                first_p.click()
        
                beg_date = browser.find_element_by_id('date_start')
                
                n = 1
                while n<=14:
                    beg_date.send_keys(Keys.BACK_SPACE)
                    n = n+1
                
                if month<10:
                    beg_date.send_keys(Keys.BACK_SPACE, '01/0'+str(month)+'/'+str(year))
                else:
                    beg_date.send_keys(Keys.BACK_SPACE, '01/'+str(month)+'/'+str(year))
        
                end_date = browser.find_element_by_id('date_end')
        
                n = 1
                while n<=14:
                    end_date.send_keys(Keys.BACK_SPACE)
                    n = n+1
    
        
                if month<10:
                    end_date.send_keys(Keys.BACK_SPACE, str(date_end)+'/0'+str(month)+'/'+str(year))
                else:
                    end_date.send_keys(Keys.BACK_SPACE, str(date_end)+'/'+str(month)+'/'+str(year))
            
        
                time.sleep(.5)
                search_elem.send_keys(Keys.ENTER)
    
            
            
                wait = WebDriverWait(browser, 60)
                element = wait.until(EC.text_to_be_present_in_element((By.ID, 'risultaticount'), "trov"))
                find_elem = browser.find_element_by_id('risultaticount')
                find_u = find_elem.text
                find_s = find_u.encode("utf-8")
                if find_s == 'Nessun documento trovato':
                    count_burocrazia = 0
                else:
                    count = ''
                    k = 9
                    while find_s[k]!=' ':
                        count = count + find_s[k]
                        k = k + 1
                        count_tit_1_p = int(count)
            
                back = browser.find_element_by_id('home')
                back.click()
                
        
                time.sleep(.5)                
                page = browser.find_element_by_xpath("//div[@id='mainsearch']/div/div[1]/table/body/tr/td[2]/div[2]/table/body/tr[3]/td[2]/div")
                page.click()
                first_p = browser.find_element_by_xpath("//div[@id='mainsearch']/div/div[1]/table/body/tr/td[2]/div[2]/table/body/tr[3]/td[2]/div/div/ul/li[1]")
                first_p.click()
                
                
        
            """UPDATE DATASET"""
            b1.cell(row = row_n, column = 1).value = year
            b1.cell(row = row_n, column = 2).value = month
            b1.cell(row = row_n, column = 3).value = count_burocrazia
            b1.cell(row = row_n, column = 4).value = count_tit
            b1.cell(row = row_n, column = 3).value = count_burocrazia_1_p
            b1.cell(row = row_n, column = 4).value = count_tit_1_p
            wb.save('Word_Count.xlsx')
        
        
            row_n += 1
            month += 1
            
        year += 1
        
except:
    server = smtplib.SMTP('smtp.gmail.com', 587)
    server.starttls()
    server.login("cmarangon94@gmail.com", "PASSWORD")
 
    msg = "An Exception Occurred!"
    server.sendmail("cmarangon94@gmail.com", "claudia.marangon@studbocconi.it", msg)
    server.quit()

    





