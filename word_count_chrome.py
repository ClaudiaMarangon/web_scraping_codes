#!/usr/bin/env python2
# -*- coding: utf-8 -*-
"""
Created on Fri Feb 23 13:58:08 2018

@author: Claudia Marangon & Philipp Steininger
"""
import os
import requests
import openpyxl
import smtplib
import time

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
os.environ["PATH"] += "C:\Users\1715481\Anaconda2"
chromedriver = '/usr/local/bin/chromedriver'
"""OPEN EXCEL FILE"""
wb = openpyxl.load_workbook('Word_Count.xlsx')
b1 = wb['Burocratico']
b1['A1']= 'Year'
b1['B1']= 'Month'
b1['C1']= 'word_freq'
b1['D1']= 'word_freq_1_p'
wb.save('Word_Count.xlsx')
"""SELECT TIME INFO + SEARCH FOR WORD"""
browser = webdriver.Chrome(executable_path=r"chromedriver.exe")



def extractInfo(year, month, row_n):
    bis1 = year%4 # leap year if bis1==0
    
    # define the end dates:        
    if month== 4 or month == 6 or month == 9 or month == 11:
        date_end = 30
    elif month == 2 and (bis1==0 and year!=1900):
        date_end = 29
    elif month == 2 and (bis1!=0 or year==1900):
        date_end = 28
    else:
        date_end = 31
    
    
    
    try:
        
# =============================================================================
#         First, we search for the total number of times the word appears:
# =============================================================================
        
        browser.get('http://archivio.corriere.it')
        wait = WebDriverWait(browser, 60)
        element = wait.until(EC.element_to_be_clickable((By.ID, 'searchLanding')))
        
        guided_s = browser.find_element_by_class_name('search-bottom-guidata')
        guided_s.click()
        
        element = wait.until(EC.element_to_be_clickable((By.ID, 'modalSearch')))
        
        search_elem = browser.find_element_by_id('modalSearch')
        search_elem.send_keys('burocratico')
        
        beg_date = browser.find_element_by_id('datepicker-from')
        if month<10:
            beg_date.send_keys(Keys.BACK_SPACE, '010'+str(month)+str(year))
        else:
            beg_date.send_keys(Keys.BACK_SPACE, '01'+str(month)+str(year))
        
        end_date = browser.find_element_by_id('datepicker-to')
    
        n = 1
        while n<=14:
            end_date.send_keys(Keys.BACK_SPACE)
            n = n+1
    
        
        if month<10:
            end_date.send_keys(Keys.BACK_SPACE, str(date_end)+'0'+str(month)+str(year))
        else:
            end_date.send_keys(Keys.BACK_SPACE, str(date_end)+str(month)+str(year))
            
        testata = browser.find_element_by_xpath("//div[@class='col-sm-8 col-md-9 form-group']/div[1]/span[2]")
        testata.click()
        cds = browser.find_element_by_xpath("//div[@class='col-sm-8 col-md-9 form-group']/div[1]/div/ul/li[1]")
        cds.click()
            
        
        time.sleep(.5)
        search_elem.send_keys(Keys.ENTER)
        element = wait.until(EC.text_to_be_present_in_element((By.ID, 'tatal_res'), "burocratico"))
        find_elem = browser.find_element_by_id('tatal_res')
        find_u = find_elem.text
        find_s = find_u.encode("utf-8")
        count = ''
        k = 12
        while find_u[k] != ' ': # orig: find_s - it does not work for the encoded find_s...
            count = count + find_u[k]
            k += 1
            count_burocrazia = int(count)
        
        time.sleep(.5)
        
        
# =============================================================================
#         Then, we check for the number of times the word appears on the first page:
# =============================================================================
        
        if count_burocrazia == 0:
            count_burocrazia_1_p = 0
        else:
            browser.get('http://archivio.corriere.it')
            element = wait.until(EC.element_to_be_clickable((By.ID, 'searchLanding')))
        
            guided_s = browser.find_element_by_class_name('search-bottom-guidata')
            guided_s.click()
        
            element = wait.until(EC.element_to_be_clickable((By.ID, 'modalSearch')))
        
            search_elem = browser.find_element_by_id('modalSearch')
            search_elem.send_keys('burocratico')
        
            beg_date = browser.find_element_by_id('datepicker-from')
            if month<10:
                beg_date.send_keys(Keys.BACK_SPACE, '010'+str(month)+str(year))
            else:
                beg_date.send_keys(Keys.BACK_SPACE, '01'+str(month)+str(year))
        
            end_date = browser.find_element_by_id('datepicker-to')
        
            n = 1
            while n<=14:
                end_date.send_keys(Keys.BACK_SPACE)
                n = n+1
    
        
            if month<10:
                end_date.send_keys(Keys.BACK_SPACE, str(date_end)+'0'+str(month)+str(year))
            else:
                end_date.send_keys(Keys.BACK_SPACE, str(date_end)+str(month)+str(year))
              
            from_p = browser.find_element_by_id('fromPage')
            from_p.send_keys('1')
        
            to_p = browser.find_element_by_id('toPage')
            to_p.send_keys(Keys.BACK_SPACE, '1')
            
            testata = browser.find_element_by_xpath("//div[@class='col-sm-8 col-md-9 form-group']/div[1]/span[2]")
            testata.click()
            cds = browser.find_element_by_xpath("//div[@class='col-sm-8 col-md-9 form-group']/div[1]/div/ul/li[1]")
            cds.click()
            
        
            time.sleep(.5)
            search_elem.send_keys(Keys.ENTER)
    
            
            
            element = wait.until(EC.text_to_be_present_in_element((By.ID, 'tatal_res'), "burocratico"))
            find_elem = browser.find_element_by_id('tatal_res')
            find_u = find_elem.text
            find_s = find_u.encode("utf-8")
            count = ''
            k = 12
            while find_u[k] !=' ': # again, changed from find_s
                count = count + find_u[k] # again, changed from find_s
                k = k + 1
            
            count_burocrazia_1_p = int(count)



        print(str(year) + '/' + str(month) + ': ' + str(count_burocrazia) + ' & ' + str(count_burocrazia_1_p))


# =============================================================================
#         Update Dataset:
# =============================================================================

        b1.cell(row = row_n, column = 1).value = year
        b1.cell(row = row_n, column = 2).value = month
        b1.cell(row = row_n, column = 3).value = count_burocrazia
        b1.cell(row = row_n, column = 4).value = count_burocrazia_1_p
        wb.save('Word_Count.xlsx')


   
# =============================================================================
# In the case of a timeout, we save the year & month where it occured.
# We will have to manually fix these missing datapoints if we do not want to use the   
# =============================================================================
    
    except TimeoutException:
        print('TimeoutException in ' + str(year) + '/' + str(month))        
        extractInfo(year, month, row_n)     
        b1.cell(row = row_n, column = 5).value = 'Exception occured!'
        wb.save('Word_Count.xlsx')
        pass
    
    except:
        return None
    
    
    
    
    

year = 1876
month = 3
row_n = 2
while year < 2017:
    if month > 12:
        month=1
        year+=1
    extractInfo(year, month, row_n)
    month+=1
    row_n+=1
    
    
    
    
    
    
    
