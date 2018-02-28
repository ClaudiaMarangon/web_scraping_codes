#!/usr/bin/env python2
# -*- coding: utf-8 -*-
"""
Created on Fri Feb 23 13:58:08 2018

@author: claudiamarangon
"""


import os
import requests
import openpyxl
import smtplib

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
os.environ["PATH"] += ":/Applications"

"""OPEN EXCEL FILE"""
wb = openpyxl.load_workbook('Word_Count.xlsx')
b1 = wb['Burocrazia']
b1['A1']= 'Year'
b1['B1']= 'Month'
b1['C1']= 'word_freq'
b1['D1']= 'word_freq_1_p'
wb.save('Word_Count.xlsx')
"""SELECT TIME INFO + SEARCH FOR WORD"""
year = 1876

row_n = 2

try:
    while year<=2017:
        bis1 = year%4    
        if year == 1876:
            month = 4
            month_in = '03'
        else:
            month = 1
            month_in = '01'
        while month<= 12:
            browser = webdriver.Firefox()
            browser.get('http://archivio.corriere.it')
            
            year = 1994
            month = 3
            search_elem = browser.find_element_by_id('searchLanding')
            search_elem.send_keys('burocrazia')
            if month== 4 or month == 6 or month == 9 or month == 11:
                date_end = 30
            elif month == 2 and (bis1==0 and year!=1900):
                date_end = 29
            elif month == 2 and (bis1!=0 or year==1900):
                date_end = 28
            else:
                date_end = 31
        

            beg_date = browser.find_element_by_id('datepicker-from-home')
            
            if month<10:
                beg_date.send_keys('010'+str(month)+str(year))
            else:
                beg_date.send_keys('01'+str(month)+str(year))
                
            end_date = browser.find_element_by_id('datepicker-to-home')
                    
            n = 1
            while n<=14:
                end_date.send_keys(Keys.BACK_SPACE)
                n = n+1

        
            if month<10:
                end_date.send_keys(Keys.BACK_SPACE, str(date_end)+'0'+str(month)+str(year))
            else:
                end_date.send_keys(Keys.BACK_SPACE, str(date_end)+str(month)+str(year))
              
            search_elem.send_keys(Keys.ENTER)
            """GET ELEMENT"""
            wait = WebDriverWait(browser, 10)
            element = wait.until(EC.text_to_be_present_in_element((By.ID, 'tatal_res'), "burocrazia"))
            find_elem = browser.find_element_by_id('tatal_res')
            find_u = find_elem.text
            find_s = find_u.encode("utf-8")
            count = ''
            k = 11
            while find_s[k]!=' ':
                count = count + find_s[k]
                k = k + 1
        
            count_burocrazia = int(count)
            
            browser.quit()
        
            """SEARCH FIRST PAGE ELEMENT"""
            browser = webdriver.Firefox()
            browser.get('http://archivio.corriere.it')
            
            guided_s = browser.find_element_by_class_name('search-bottom-guidata')
            guided_s.click()
        
            search_elem = browser.find_element_by_id('modalSearch')
            search_elem.send_keys('burocrazia')
        
            if month== 4 or month == 6 or month == 9 or month == 11:
                date_end = 30
            elif month == 2 and (bis1==0 and year!=1900):
                date_end = 29
            elif month == 2 and (bis1!=0 or year==1900):
                date_end = 28
            else:
                date_end = 31
        

            beg_date = browser.find_element_by_id('datepicker-from')
        
            if month<10:
                beg_date.send_keys(Keys.BACK_SPACE, '010'+str(month)+str(year))
            else:
                beg_date.send_keys(Keys.BACK_SPACE, '01'+str(month)+str(year))
        
            end_date = browser.find_element_by_id('datepicker-to')
        
            n = 1
            while n<=14:
                end_date.send_keys(Keys.BACK_SPACE)
                n = n+1

        
            if month<10:
                end_date.send_keys(Keys.BACK_SPACE, str(date_end)+'0'+str(month)+str(year))
            else:
                end_date.send_keys(Keys.BACK_SPACE, str(date_end)+str(month)+str(year))
              
            from_p = browser.find_element_by_id('fromPage')
            from_p.send_keys('1')
        
            to_p = browser.find_element_by_id('toPage')
            to_p.send_keys(Keys.BACK_SPACE, '1')
        
        
            search_elem.send_keys(Keys.ENTER)
        
            """GET FIRST PAGE ELEMENT"""
        
            wait = WebDriverWait(browser, 10)
            element = wait.until(EC.text_to_be_present_in_element((By.ID, 'tatal_res'), "burocrazia"))
            find_elem = browser.find_element_by_id('tatal_res')
            find_u = find_elem.text
            find_s = find_u.encode("utf-8")
            count = ''
            k = 11
            while find_s[k]!=' ':
                count = count + find_s[k]
                k = k + 1
        
            count_burocrazia_1_p = int(count)
        
        
            """UPDATE DATASET"""
            b1.cell(row = row_n, column = 1).value = year
            b1.cell(row = row_n, column = 2).value = month
            b1.cell(row = row_n, column = 3).value = count_burocrazia
            b1.cell(row = row_n, column = 4).value = count_burocrazia_1_p
            wb.save('Word_Count.xlsx')
        
            browser.quit()
            row_n = row_n + 1
            month = month + 1
            
        year = year + 1
except:
    server = smtplib.SMTP('smtp.gmail.com', 587)
    server.starttls()
    server.login("cmarangon94@gmail.com", "Clapi_pini13")
 
    msg = "An Exception Occurred!"
    server.sendmail("cmarangon94@gmail.com", "claudia.marangon@studbocconi.it", msg)
    server.quit()




